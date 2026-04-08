"""Dashboard API endpoints for the BDC metrics web application."""

import io
import logging
from datetime import date, timedelta

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from src.api.services.gross_sales import get_gross_sales_data
from src.api.services.redemptions import get_redemptions_data
from src.api.services.performance import get_performance_data
from src.api.services.redemption_requests import get_redemption_requests_data
from src.api.services.net_flows import get_net_flows_data

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/dashboard")


def _default_dates(
    start: str | None,
    end: str | None,
) -> tuple[date, date]:
    """Parse date params or return defaults (last 12 months)."""
    if end:
        end_date = date.fromisoformat(end + "-01") if len(end) == 7 else date.fromisoformat(end)
    else:
        end_date = date.today()

    if start:
        start_date = date.fromisoformat(start + "-01") if len(start) == 7 else date.fromisoformat(start)
    else:
        start_date = date(end_date.year - 1, end_date.month, 1)

    return start_date, end_date


@router.get("/gross-sales")
async def dashboard_gross_sales(
    start: str | None = Query(None, description="Start date (YYYY-MM or YYYY-MM-DD)"),
    end: str | None = Query(None, description="End date (YYYY-MM or YYYY-MM-DD)"),
    period: str = Query("monthly", description="monthly or quarterly"),
):
    start_date, end_date = _default_dates(start, end)
    return await get_gross_sales_data(start_date, end_date, period)


@router.get("/redemptions")
async def dashboard_redemptions(
    start: str | None = Query(None),
    end: str | None = Query(None),
    period: str = Query("monthly"),
):
    start_date, end_date = _default_dates(start, end)
    return await get_redemptions_data(start_date, end_date, period)


@router.get("/performance")
async def dashboard_performance(
    start: str | None = Query(None),
    end: str | None = Query(None),
    period: str = Query("monthly"),
):
    start_date, end_date = _default_dates(start, end)
    return await get_performance_data(start_date, end_date, period)


@router.get("/redemption-requests")
async def dashboard_redemption_requests(
    start: str | None = Query(None),
    end: str | None = Query(None),
    period: str = Query("monthly"),
):
    start_date, end_date = _default_dates(start, end)
    return await get_redemption_requests_data(start_date, end_date, period)


@router.get("/net-flows")
async def dashboard_net_flows(
    start: str | None = Query(None),
    end: str | None = Query(None),
    period: str = Query("monthly"),
):
    start_date, end_date = _default_dates(start, end)
    return await get_net_flows_data(start_date, end_date, period)


TAB_CONFIG = [
    ("Gross Sales", get_gross_sales_data, "monthly"),
    ("Redemptions", get_redemptions_data, "quarterly"),
    ("Performance", get_performance_data, "monthly"),
    ("Redemption Requests", get_redemption_requests_data, "quarterly"),
    ("Net Flows", get_net_flows_data, "quarterly"),
]


@router.get("/export")
async def export_xlsx(
    start: str | None = Query(None),
    end: str | None = Query(None),
    period: str = Query("monthly"),
):
    """Export all tabs as an XLSX file with one sheet per tab."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    start_date, end_date = _default_dates(start, end)

    wb = Workbook()
    wb.remove(wb.active)

    MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

    def _fmt_date(ds: str) -> str:
        """'2025-03-31' -> 'Mar 2025'"""
        d = date.fromisoformat(ds)
        return f"{MONTHS[d.month - 1]} {d.year}"

    header_font = Font(name="Calibri", bold=True, size=10)
    header_fill = PatternFill(start_color="F2F3F5", end_color="F2F3F5", fill_type="solid")
    total_font = Font(name="Calibri", bold=True, size=10)
    total_fill = PatternFill(start_color="F2F3F5", end_color="F2F3F5", fill_type="solid")
    bank_font = Font(name="Calibri", bold=True, size=10, color="333333")
    body_font = Font(name="Calibri", size=10)
    pct_fmt = '0.0%'
    pct0_fmt = '0%'
    cur_fmt = '$#,##0'
    num_fmt = '#,##0.0'

    for tab_name, service_fn, default_period in TAB_CONFIG:
        p = default_period if default_period == "quarterly" else period
        data = await service_fn(start_date, end_date, p)
        funds = data.get("funds", [])
        banks = data.get("banks", [])
        fund_cols = list(funds) + ["Total"]

        ws = wb.create_sheet(title=tab_name[:31])
        row_num = 1

        for bank in banks:
            # Bank header
            ws.cell(row=row_num, column=1, value=bank["name"]).font = bank_font
            row_num += 1

            bank_rows = bank.get("rows", [])
            fmt = bank.get("format", "")

            # Column headers: Fund | date1 | date2 | ...
            ws.cell(row=row_num, column=1, value="").font = header_font
            for ci, brow in enumerate(bank_rows, start=2):
                cell = ws.cell(row=row_num, column=ci, value=_fmt_date(brow["date"]))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            row_num += 1

            # Data rows
            for fund in fund_cols:
                is_total = fund == "Total"
                c1 = ws.cell(row=row_num, column=1, value=fund)
                c1.font = total_font if is_total else body_font
                if is_total:
                    c1.fill = total_fill

                for ci, brow in enumerate(bank_rows, start=2):
                    val = brow.get(fund)
                    cell = ws.cell(row=row_num, column=ci)
                    if is_total:
                        cell.font = total_font
                        cell.fill = total_fill
                    else:
                        cell.font = body_font
                    cell.alignment = Alignment(horizontal="right")

                    if val is None:
                        cell.value = "-"
                    elif val == "N/A":
                        cell.value = "N/A"
                    elif fmt in ("percent", "percent1"):
                        cell.value = float(val)
                        cell.number_format = pct_fmt if fmt == "percent1" else pct0_fmt
                    elif fmt == "currency":
                        cell.value = round(float(val))
                        cell.number_format = cur_fmt
                    elif fmt == "number":
                        cell.value = float(val)
                        cell.number_format = num_fmt
                    else:
                        cell.value = val

                row_num += 1

            row_num += 1  # blank row between banks

        # Auto-width columns
        for col_idx in range(1, len(bank_rows) + 2 if banks and banks[0].get("rows") else 2):
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = 14 if col_idx > 1 else 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"bdc_metrics_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
