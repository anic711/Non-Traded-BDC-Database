#!/usr/bin/env python3
"""Export BDC metrics to Excel with calculated fields.

Calculations:
1. Monthly performance by class = (ΔNAV + distribution) / NAV(t-1)
2. Monthly gross sales = Σ(Δshares_issued × NAV(t-1)) across classes
3. Monthly gross sales (alt) = Σ(Δtotal_consideration) across classes
4. Monthly shares outstanding = interpolated from quarterly actuals
"""

import sqlite3
from collections import defaultdict
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers

DB_PATH = "bdc_metrics.db"
OUTPUT_PATH = "bdc_metrics_export.xlsx"

header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
pct_fmt = '0.00%'
money_fmt = '#,##0'
nav_fmt = '#,##0.0000'
shares_fmt = '#,##0'


def style_header(ws):
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill


def auto_width(ws, min_width=12):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, 30))


def load_data(conn):
    """Load all raw data from the database."""
    c = conn.cursor()

    data = {}
    c.execute("SELECT id, ticker FROM funds ORDER BY ticker")
    funds = c.fetchall()

    for fund_id, ticker in funds:
        d = {}

        # NAV per share: {date: {class: nav}}
        c.execute("""SELECT as_of_date, share_class, nav_per_share
                     FROM nav_per_share WHERE fund_id = ? ORDER BY as_of_date""", (fund_id,))
        nav = defaultdict(dict)
        for dt, cls, val in c.fetchall():
            nav[dt][cls] = float(val) if val else None
        d['nav'] = dict(nav)

        # Distributions: {date: {class: dist}}
        c.execute("""SELECT as_of_date, share_class, distribution_per_share
                     FROM distributions WHERE fund_id = ? ORDER BY as_of_date""", (fund_id,))
        dist = defaultdict(dict)
        for dt, cls, val in c.fetchall():
            dist[dt][cls] = float(val) if val else None
        d['dist'] = dict(dist)

        # Shares issued (cumulative): {date: {class: total_shares}}
        # Sum across offering types for each class
        c.execute("""SELECT as_of_date, share_class,
                            SUM(cumulative_shares), SUM(cumulative_consideration)
                     FROM shares_issued WHERE fund_id = ?
                     GROUP BY as_of_date, share_class
                     ORDER BY as_of_date""", (fund_id,))
        issued_shares = defaultdict(dict)
        issued_consid = defaultdict(dict)
        for dt, cls, sh, co in c.fetchall():
            issued_shares[dt][cls] = float(sh) if sh else 0
            issued_consid[dt][cls] = float(co) if co else 0
        d['issued_shares'] = dict(issued_shares)
        d['issued_consid'] = dict(issued_consid)

        # Redemptions: {date: {tendered, shares, value}}
        c.execute("""SELECT as_of_date, shares_tendered, shares_redeemed, value_redeemed
                     FROM redemptions WHERE fund_id = ? ORDER BY as_of_date""", (fund_id,))
        redemptions = {}
        for dt, tend, sh, val in c.fetchall():
            redemptions[dt] = {
                'tendered': float(tend) if tend else 0,
                'shares': float(sh) if sh else 0,
                'value': float(val) if val else 0,
            }
        d['redemptions'] = redemptions

        # Total NAV: {date: total_nav}
        c.execute("""SELECT as_of_date, total_nav
                     FROM total_nav WHERE fund_id = ? ORDER BY as_of_date""", (fund_id,))
        d['total_nav'] = {dt: float(val) for dt, val in c.fetchall() if val}

        # Shares outstanding: {date: total_shares}
        c.execute("""SELECT as_of_date, total_shares_outstanding
                     FROM shares_outstanding WHERE fund_id = ? ORDER BY as_of_date""", (fund_id,))
        d['shares_outstanding'] = {dt: float(val) for dt, val in c.fetchall() if val}

        data[ticker] = d

    return data


def get_nav_classes(nav_data):
    """Get sorted list of share classes from NAV data."""
    classes = set()
    for by_class in nav_data.values():
        classes.update(by_class.keys())
    return sorted(classes)


def calc_performance(nav_data, dist_data):
    """Calculate monthly performance by class.

    Performance = (NAV_t - NAV_{t-1} + dist_t) / NAV_{t-1}
    """
    classes = get_nav_classes(nav_data)
    dates = sorted(nav_data.keys())
    results = []

    for i in range(1, len(dates)):
        dt = dates[i]
        dt_prev = dates[i - 1]
        row = {'date': dt}
        for cls in classes:
            nav_t = nav_data[dt].get(cls)
            nav_prev = nav_data[dt_prev].get(cls)
            dist_t = dist_data.get(dt, {}).get(cls, 0) or 0
            if nav_t is not None and nav_prev is not None and nav_prev > 0:
                row[cls] = (nav_t - nav_prev + dist_t) / nav_prev
            else:
                row[cls] = None
        results.append(row)

    return results, classes


def calc_gross_sales_nav(nav_data, issued_shares_data):
    """Calculate monthly gross sales = Σ(Δshares_issued × NAV(t-1)) across classes.

    Monthly shares issued = cumulative_t - cumulative_{t-1}
    Gross sales for class = monthly_shares_issued × NAV_per_share_{t-1}
    Total gross sales = sum across classes
    """
    issued_dates = sorted(issued_shares_data.keys())
    nav_dates = sorted(nav_data.keys())
    classes = get_nav_classes(nav_data)
    results = []

    for i in range(1, len(issued_dates)):
        dt = issued_dates[i]
        dt_prev = issued_dates[i - 1]

        # Find the NAV date closest to (but not after) dt_prev
        nav_prev = None
        for nd in reversed(nav_dates):
            if nd <= dt_prev:
                nav_prev = nav_data[nd]
                break
        if not nav_prev:
            continue

        row = {'date': dt, 'total': 0}
        for cls in classes:
            cum_t = issued_shares_data[dt].get(cls, 0)
            cum_prev = issued_shares_data[dt_prev].get(cls, 0)
            delta_shares = cum_t - cum_prev
            nav_price = nav_prev.get(cls)

            if delta_shares > 0 and nav_price and nav_price > 0:
                sales = delta_shares * nav_price
                row[cls] = sales
                row['total'] += sales
            else:
                row[cls] = 0 if delta_shares == 0 else None
                if row[cls] == 0:
                    row['total'] += 0

        results.append(row)

    return results, classes


def calc_gross_sales_consideration(issued_consid_data):
    """Calculate monthly gross sales (alt) = Σ(Δtotal_consideration) across classes."""
    dates = sorted(issued_consid_data.keys())
    classes = set()
    for by_class in issued_consid_data.values():
        classes.update(by_class.keys())
    classes = sorted(classes)

    results = []
    for i in range(1, len(dates)):
        dt = dates[i]
        dt_prev = dates[i - 1]
        row = {'date': dt, 'total': 0}
        for cls in classes:
            cum_t = issued_consid_data[dt].get(cls, 0)
            cum_prev = issued_consid_data[dt_prev].get(cls, 0)
            delta = cum_t - cum_prev
            row[cls] = delta
            row['total'] += delta
        results.append(row)

    return results, classes


def is_quarter_end(dt_str):
    """Check if a date string is a quarter-end month."""
    d = date.fromisoformat(dt_str)
    return d.month in (3, 6, 9, 12)


def get_prior_quarter_end(dt_str):
    """Get the most recent quarter-end date on or before the given date."""
    d = date.fromisoformat(dt_str)
    # If this is a quarter-end month, the prior quarter end is the previous one
    month = d.month
    year = d.year
    # Map to the most recent quarter-end that has passed
    qe_months = [12, 9, 6, 3]
    for qm in qe_months:
        qe_year = year if qm <= month else year - 1
        if qm < month or (qm == month):
            # This quarter end is on or before our month
            return f"{qe_year}-{qm:02d}-{[0,31,28,31,30,31,30,31,31,30,31,30,31][qm]:02d}"
    return None


def calc_monthly_shares_outstanding(shares_outstanding, issued_shares, redemptions):
    """Calculate monthly shares outstanding.

    Quarter-end months: use reported shares_outstanding
    Other months: last_quarter_outstanding + QTD_shares_issued - QTD_shares_redeemed
    """
    # Build a timeline of all months from issued_shares dates
    issued_dates = sorted(issued_shares.keys())
    if not issued_dates:
        return []

    # Map quarter-end dates to reported shares outstanding
    # The shares_outstanding dates may not be exact quarter ends, so find closest
    qe_reported = {}
    so_dates = sorted(shares_outstanding.keys())

    for so_date in so_dates:
        d = date.fromisoformat(so_date)
        # Check if this is near a quarter end (within 15 days)
        for qm in [3, 6, 9, 12]:
            import calendar
            last_day = calendar.monthrange(d.year, qm)[1]
            qe = date(d.year, qm, last_day)
            if abs((d - qe).days) <= 15:
                qe_str = qe.isoformat()
                qe_reported[qe_str] = shares_outstanding[so_date]
                break

    # Also build total cumulative shares issued per date (sum across classes)
    total_issued = {}
    for dt, by_class in issued_shares.items():
        total_issued[dt] = sum(by_class.values())

    # Build redemptions lookup (quarterly)
    # redemptions dict: {date_str: {shares: X, value: Y}}

    results = []
    for dt in issued_dates:
        d = date.fromisoformat(dt)

        # Check if we have a reported value for this quarter end
        # Try exact match to common quarter-end dates
        qe_match = None
        for qm in [3, 6, 9, 12]:
            import calendar
            last_day = calendar.monthrange(d.year, qm)[1]
            qe = date(d.year, qm, last_day)
            if abs((d - qe).days) <= 5:
                qe_str = qe.isoformat()
                if qe_str in qe_reported:
                    qe_match = qe_reported[qe_str]
                break

        if qe_match is not None:
            results.append({'date': dt, 'shares_outstanding': qe_match, 'source': 'reported'})
            continue

        # Find the most recent quarter-end with reported data
        best_qe = None
        best_qe_val = None
        for qe_dt in sorted(qe_reported.keys()):
            if qe_dt <= dt:
                best_qe = qe_dt
                best_qe_val = qe_reported[qe_dt]

        if best_qe_val is None:
            results.append({'date': dt, 'shares_outstanding': None, 'source': 'no_base'})
            continue

        # Get cumulative issued at quarter end and current month
        # Find the issued date closest to the quarter end
        cum_issued_qe = None
        for id_dt in sorted(total_issued.keys()):
            d_id = date.fromisoformat(id_dt)
            d_qe = date.fromisoformat(best_qe)
            if abs((d_id - d_qe).days) <= 5:
                cum_issued_qe = total_issued[id_dt]
                break
            elif id_dt <= best_qe:
                cum_issued_qe = total_issued[id_dt]

        cum_issued_now = total_issued.get(dt, None)

        if cum_issued_qe is None or cum_issued_now is None:
            results.append({'date': dt, 'shares_outstanding': None, 'source': 'no_issued'})
            continue

        delta_issued = cum_issued_now - cum_issued_qe

        # Check for any redemptions between quarter end and now
        delta_redeemed = 0
        for r_dt, r_data in redemptions.items():
            if best_qe < r_dt <= dt:
                delta_redeemed += r_data.get('shares', 0)

        est_outstanding = best_qe_val + delta_issued - delta_redeemed
        results.append({'date': dt, 'shares_outstanding': est_outstanding, 'source': 'estimated'})

    return results


def write_raw_sheets(wb, ticker, d):
    """Write the raw data sheets for a BDC."""
    fund_data = d

    # NAV Per Share
    ws = wb.create_sheet(f"{ticker} - NAV")
    classes = get_nav_classes(fund_data['nav'])
    ws.append(["Date"] + classes)
    style_header(ws)
    for dt in sorted(fund_data['nav']):
        row = [dt] + [fund_data['nav'][dt].get(c) for c in classes]
        ws.append(row)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is not None:
                cell.number_format = nav_fmt
    auto_width(ws)

    # Distributions
    ws = wb.create_sheet(f"{ticker} - Distributions")
    dist_classes = set()
    for by_c in fund_data['dist'].values():
        dist_classes.update(by_c.keys())
    dist_classes = sorted(dist_classes)
    ws.append(["Date"] + dist_classes)
    style_header(ws)
    for dt in sorted(fund_data['dist']):
        row = [dt] + [fund_data['dist'][dt].get(c) for c in dist_classes]
        ws.append(row)
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if cell.value is not None:
                cell.number_format = nav_fmt
    auto_width(ws)

    # Shares Issued (from DB directly — need offering type detail)
    conn2 = sqlite3.connect(DB_PATH)
    c2 = conn2.cursor()
    c2.execute("SELECT id FROM funds WHERE ticker = ?", (ticker,))
    fund_id = c2.fetchone()[0]
    ws = wb.create_sheet(f"{ticker} - Shares Issued")
    ws.append(["Date", "Share Class", "Offering Type", "Cumulative Shares", "Cumulative Consideration"])
    style_header(ws)
    c2.execute("""SELECT as_of_date, share_class, offering_type, cumulative_shares, cumulative_consideration
                  FROM shares_issued WHERE fund_id = ? ORDER BY as_of_date, share_class, offering_type""", (fund_id,))
    for dt, cls, off, sh, co in c2.fetchall():
        ws.append([dt, cls, off, float(sh) if sh else 0, float(co) if co else 0])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = shares_fmt
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = money_fmt
    auto_width(ws)
    conn2.close()

    # Redemptions
    ws = wb.create_sheet(f"{ticker} - Redemptions")
    ws.append(["Date", "Shares Tendered", "Shares Redeemed", "Value Redeemed"])
    style_header(ws)
    for dt in sorted(fund_data['redemptions']):
        r = fund_data['redemptions'][dt]
        ws.append([dt, r['tendered'], r['shares'], r['value']])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = shares_fmt
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = money_fmt
    auto_width(ws)

    # Total NAV
    ws = wb.create_sheet(f"{ticker} - Total NAV")
    ws.append(["Date", "Total NAV"])
    style_header(ws)
    for dt in sorted(fund_data['total_nav']):
        ws.append([dt, fund_data['total_nav'][dt]])
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = money_fmt
    auto_width(ws)

    # Shares Outstanding (reported)
    ws = wb.create_sheet(f"{ticker} - Shares Outstanding")
    ws.append(["Date", "Total Common Shares Outstanding"])
    style_header(ws)
    for dt in sorted(fund_data['shares_outstanding']):
        ws.append([dt, fund_data['shares_outstanding'][dt]])
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = shares_fmt
    auto_width(ws)


def write_calc_sheets(wb, ticker, d):
    """Write the calculated sheets for a BDC."""

    # 1. Monthly Performance
    perf, classes = calc_performance(d['nav'], d['dist'])
    ws = wb.create_sheet(f"{ticker} - Performance")
    ws.append(["Date"] + classes)
    style_header(ws)
    for row in perf:
        ws.append([row['date']] + [row.get(c) for c in classes])
    for r in ws.iter_rows(min_row=2, min_col=2):
        for cell in r:
            if cell.value is not None:
                cell.number_format = pct_fmt
    auto_width(ws)

    # 2. Monthly Gross Sales (NAV method)
    gs_nav, classes = calc_gross_sales_nav(d['nav'], d['issued_shares'])
    ws = wb.create_sheet(f"{ticker} - Gross Sales")
    ws.append(["Date"] + classes + ["Total"])
    style_header(ws)
    for row in gs_nav:
        ws.append([row['date']] + [row.get(c) for c in classes] + [row.get('total')])
    for r in ws.iter_rows(min_row=2, min_col=2):
        for cell in r:
            if cell.value is not None:
                cell.number_format = money_fmt
    auto_width(ws)

    # 3. Monthly Gross Sales (Consideration method)
    gs_con, classes = calc_gross_sales_consideration(d['issued_consid'])
    ws = wb.create_sheet(f"{ticker} - Gross Sales (Alt)")
    ws.append(["Date"] + classes + ["Total"])
    style_header(ws)
    for row in gs_con:
        ws.append([row['date']] + [row.get(c) for c in classes] + [row.get('total')])
    for r in ws.iter_rows(min_row=2, min_col=2):
        for cell in r:
            if cell.value is not None:
                cell.number_format = money_fmt
    auto_width(ws)

    # 4. Monthly Shares Outstanding
    monthly_so = calc_monthly_shares_outstanding(
        d['shares_outstanding'], d['issued_shares'], d['redemptions']
    )
    ws = wb.create_sheet(f"{ticker} - Monthly Shares Out")
    ws.append(["Date", "Shares Outstanding", "Source"])
    style_header(ws)
    for row in monthly_so:
        ws.append([row['date'],
                   row['shares_outstanding'],
                   row['source']])
    for r in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in r:
            if cell.value is not None:
                cell.number_format = shares_fmt
    auto_width(ws)


def main():
    conn = sqlite3.connect(DB_PATH)
    data = load_data(conn)
    conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    for ticker in sorted(data.keys()):
        print(f"Processing {ticker}...")
        write_raw_sheets(wb, ticker, data[ticker])
        write_calc_sheets(wb, ticker, data[ticker])

    wb.save(OUTPUT_PATH)
    print(f"\nSaved {OUTPUT_PATH} with {len(wb.sheetnames)} sheets:")
    for s in wb.sheetnames:
        print(f"  {s}")


if __name__ == "__main__":
    main()
